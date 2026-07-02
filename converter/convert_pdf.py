from __future__ import annotations

import base64
import json
import re
import sys
import tempfile
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

try:
    import fitz  # type: ignore[import-not-found]
except Exception:  # pragma: no cover - optional dependency
    fitz = None

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

LOGS: list[str] = []


def log(msg: str) -> None:
    LOGS.append(msg)
    print(msg, file=sys.stderr)


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in str(value).split("\n")]
    return "\n".join(lines).strip()


def parse_price(val: str) -> float:
    if not val:
        return 0.0
    cleaned = val.replace(",", "").replace("USD", "").replace("$", "").replace("-", "").strip()
    if not cleaned:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def find_latest_master_path(output_dir: Path) -> Path | None:
    if not output_dir.exists():
        return None
    import glob
    pattern = str(output_dir / "專案匯總管理表*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort()
    return Path(files[-1])


def get_display_length(s: str) -> int:
    if not s:
        return 0
    return sum(2 if ord(c) > 127 else 1 for c in s)


def style_range(ws, cell_range: str, border: Border = None, fill: PatternFill = None, font: Font = None, alignment: Alignment = None):
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border

def style_merged_range_borders(ws, cell_range: str, border_side: Side):
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            left = border_side if c == min_col else None
            right = border_side if c == max_col else None
            top = border_side if r == min_row else None
            bottom = border_side if r == max_row else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def copy_sheet_to_workbook(src_ws, target_wb, target_name):
    # If a sheet with the same name already exists in target_wb, remove it first
    if target_name in target_wb.sheetnames:
        target_wb.remove(target_wb[target_name])
        
    dest_ws = target_wb.create_sheet(title=target_name)
    
    # Copy values, formulas and formatting cell-by-cell
    for r in range(1, src_ws.max_row + 1):
        # Copy row height
        if r in src_ws.row_dimensions:
            dest_ws.row_dimensions[r].height = src_ws.row_dimensions[r].height
            
        for c in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(r, c)
            dest_cell = dest_ws.cell(r, c)
            
            dest_cell.value = src_cell.value
            
            # Copy styling
            if src_cell.has_style:
                import copy
                dest_cell.font = copy.copy(src_cell.font) if src_cell.font else None
                dest_cell.fill = copy.copy(src_cell.fill) if src_cell.fill else None
                dest_cell.border = copy.copy(src_cell.border) if src_cell.border else None
                dest_cell.alignment = copy.copy(src_cell.alignment) if src_cell.alignment else None
                dest_cell.number_format = src_cell.number_format
                
    # Copy merged ranges
    for merged_range in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged_range))
        
    # Copy column widths
    for col_idx in range(1, src_ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in src_ws.column_dimensions:
            dest_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
            
    # Copy logo images
    import copy
    for img in src_ws._images:
        try:
            new_img = copy.copy(img)
            new_img.anchor = img.anchor
            if img.width and img.height:
                new_img.width = img.width
                new_img.height = img.height
            dest_ws.add_image(new_img)
        except Exception as e:
            log(f"Error copying image: {e}")
            
    # Copy freeze panes
    dest_ws.freeze_panes = src_ws.freeze_panes

def apply_outer_borders(ws, start_row, end_row, start_col, end_col, divider_row=None):
    thin_side = Side(style="thin", color="000000")
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            current_border = cell.border
            
            left_border = thin_side if c == start_col else (current_border.left if current_border else None)
            right_border = thin_side if c == end_col else (current_border.right if current_border else None)
            top_border = thin_side if r == start_row else (current_border.top if current_border else None)
            
            is_bottom = (r == end_row) or (divider_row is not None and r == divider_row)
            bottom_border = thin_side if is_bottom else (current_border.bottom if current_border else None)
            
            if divider_row is not None and r == divider_row + 1:
                top_border = thin_side
                
            cell.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)

def parse_customer_info(text: str) -> dict[str, str]:
    info = {"customer": "", "contact": "", "phone": "", "date": "", "project": "", "mail": "", "raw_text": text}
    if not text:
        return info
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    for line in lines:
        cust_match = re.search(r'(?:客戶|客户)\s*:\s*(.*?)(?=(?:聯繫|聯絡人|联系|報價日期|报价日期|電話|电话|Mail|mail|$))', line)
        if cust_match:
            info["customer"] = cust_match.group(1).strip()
            
        date_match = re.search(r'(?:報價日期|报价日期|日期)\s*:\s*(\d{4}[./-]\d{2}[./-]\d{2})', line)
        if date_match:
            info["date"] = date_match.group(1).strip()
            
        contact_match = re.search(r'(?:聯繫|联絡人|联系)\s*:\s*(.*?)(?=(?:客戶|客户|報價日期|报价日期|電話|电话|Mail|mail|$))', line)
        if contact_match:
            info["contact"] = contact_match.group(1).strip()
            
        if "專案" in line and ":" in line:
            val = line.split(":", 1)[1].strip()
            if not any(kw in val for kw in ("系統", "系統", "報價", "报价")):
                info["contact"] = val
                
        phone_match = re.search(r'(?:電話|电话).+?(\+?[\d\s-]{9,})', line)
        if phone_match:
            info["phone"] = phone_match.group(1).strip()
        else:
            phone_match = re.search(r'(?:電話|电话)\s*:\s*([\d\s+-]*)', line)
            if phone_match and phone_match.group(1).strip():
                info["phone"] = phone_match.group(1).strip()
                
        mail_match = re.search(r'(?:Mail|mail|Email|email)\s*:\s*(.*?)(?=(?:客戶|客户|報價日期|报价日期|電話|电话|$))', line)
        if mail_match:
            info["mail"] = mail_match.group(1).strip()
            
        if re.match(r'^\+?[\d-]+$', line):
            info["phone"] = line
            
    # Project Title extraction:
    # First check if there is a line in the block containing "液冷", "系統", "系統", "報價", "报价" but without labels
    for line in lines:
        clean_line = line.strip()
        has_label = any(clean_line.startswith(pref) for pref in ("客戶", "客户", "聯繫", "联系", "聯絡人", "Mail", "mail", "電話", "电话", "日期", "報價日期", "报价日期"))
        if not has_label:
            if any(kw in clean_line for kw in ("液冷", "系統", "系统", "報價", "报价", "數據中心", "数据中心")):
                info["project"] = clean_line
                break
                
    if not info["project"]:
        # Fallback to general regex
        project_match = re.search(r'([^\s\n]*(?:報價|报价|系統|系统)(?!日期)[^\s\n]*)', text)
        if project_match:
            info["project"] = project_match.group(1).strip()
            info["project"] = re.sub(r'\s*No\.\s*\d+', '', info["project"])
            
    return info

def split_remarks(text: str) -> list[str]:
    if not text:
        return []
    raw_lines = [line.strip() for line in text.split("\n") if line.strip()]
    final_lines = []
    for line in raw_lines:
        matches = list(re.finditer(r'(\d+\))', line))
        if not matches:
            final_lines.append(line)
            continue
        first_part = line[:matches[0].start()].strip()
        if first_part:
            final_lines.append(first_part)
        for i in range(len(matches)):
            start_pos = matches[i].start()
            end_pos = matches[i+1].start() if i + 1 < len(matches) else len(line)
            part = line[start_pos:end_pos].strip()
            if part:
                final_lines.append(part)
    return final_lines


def extract_beehe_document(pdf_path: Path, target_pages: set[int] | None = None) -> list[dict[str, object]]:
    log(f"Extracting Beehe tables from document: {pdf_path.name}")
    tables = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            if target_pages is not None and page_index not in target_pages:
                continue
            
            find_tables = page.find_tables() or []
            if not find_tables:
                continue
                
            for table_idx, t in enumerate(find_tables, start=1):
                raw_table = t.extract()
                grid = [[clean_cell(cell) for cell in row] for row in raw_table]
                
                header_idx = -1
                for idx, r in enumerate(grid):
                    row_joined = " ".join(r).lower()
                    if "no" in row_joined and ("專案" in row_joined or "項目" in row_joined or "內容" in row_joined):
                        header_idx = idx
                        break
                        
                if header_idx == -1:
                    log(f"Page {page_index} Table {table_idx}: Header row with 'No.' not found.")
                    continue
                    
                is_split_header = False
                if header_idx > 0:
                    prev_row_vals = grid[header_idx - 1]
                    if not prev_row_vals[0].strip() and any(prev_row_vals):
                        is_split_header = True
                
                if is_split_header:
                    top_y = t.rows[header_idx - 1].bbox[1]
                else:
                    top_y = t.rows[header_idx].bbox[1]
                    
                # Parse customer info
                cust_cells = [cell.strip() for cell in grid[1] if cell] if len(grid) > 1 else []
                cust_text = "\n".join(cust_cells)
                customer_info = parse_customer_info(cust_text)
                
                headers = []
                for col_idx in range(len(grid[header_idx])):
                    parts = []
                    for r_idx in (header_idx - 1, header_idx, header_idx + 1):
                        if 0 <= r_idx < len(grid):
                            if grid[r_idx] and re.match(r'^\d+$', str(grid[r_idx][0]).strip()):
                                continue
                            val = grid[r_idx][col_idx].strip()
                            if val and val not in parts:
                                parts.append(val)
                    headers.append(" ".join(parts))
                    
                col_no = 0
                col_item = 1
                col_desc = 2
                col_price = 3
                col_warranty = 4
                col_qty = 5
                col_unit = 6
                col_total = 7
                
                for i, h in enumerate(headers):
                    h_clean = h.lower().replace(" ", "")
                    if "no" in h_clean:
                        col_no = i
                    elif "內容" in h_clean:
                        col_desc = i
                    elif "專案" in h_clean or "項目" in h_clean:
                        col_item = i
                    elif "單價(usd)/台" in h_clean or "保固單價" in h_clean or (i == 4 and "保固" in h_clean):
                        col_warranty = i
                    elif "單價" in h_clean:
                        col_price = i
                    elif "qty" in h_clean or "數量" in h_clean:
                        col_qty = i
                        if i + 1 < len(headers) and not headers[i + 1]:
                            col_unit = i + 1
                    elif "總計" in h_clean or "ttl" in h_clean or "total" in h_clean:
                        col_total = i
                
                has_remark_col = False
                remark_col_idx = -1
                for i, h in enumerate(headers):
                    if "備註" in h or "remark" in h.lower():
                        has_remark_col = True
                        remark_col_idx = i
                        break
                
                col_order = [col_no, col_item, col_desc, col_qty, col_unit, col_price, col_warranty, col_total]
                if has_remark_col:
                    col_order.append(remark_col_idx)
                    
                new_headers = [headers[idx] for idx in col_order]
                
                data_rows = []
                total_row_raw = None
                total_row_idx = -1
                
                for r_idx, r in enumerate(grid[header_idx + 1:], start=header_idx + 1):
                    if not r or not any(c.strip() for c in r if c is not None):
                        continue
                    first_cell = r[0].strip()
                    if re.match(r'^\d+$', first_cell):
                        data_rows.append(r)
                    elif first_cell.lower().startswith("ttl") or first_cell.lower().startswith("total") or re.search(r'\b(ttl|total)\b', " ".join(r).lower()):
                        total_row_raw = r
                        total_row_idx = r_idx
                
                if total_row_idx != -1:
                    bottom_y = t.rows[total_row_idx].bbox[3]
                else:
                    bottom_y = t.rows[-1].bbox[3]
                    
                # Parse remarks
                remarks_text_parts = []
                if total_row_idx != -1:
                    for r_rem in grid[total_row_idx + 1:]:
                        if r_rem and r_rem[0]:
                            remarks_text_parts.append(r_rem[0])
                remarks_text = "\n".join(remarks_text_parts)
                remarks_lines = split_remarks(remarks_text)
                
                is_new_version = any("項目" in h for h in headers)
                
                r_col_no = 0
                r_col_item = 1
                r_col_desc = 2
                r_col_qty = 3
                r_col_unit = 4
                r_col_price = 5
                r_col_warranty = 6
                r_col_total = 7
                
                reordered_data_rows = []
                for r in data_rows:
                    this_col_order = [col_no, col_item, col_desc, col_qty, col_unit, col_price, col_warranty, col_total]
                    if has_remark_col:
                        this_col_order.append(remark_col_idx)
                        
                    reordered_row = [r[idx] if idx < len(r) else "" for idx in this_col_order]
                    if len(reordered_row) > r_col_item:
                        reordered_row[r_col_item] = reordered_row[r_col_item].replace("比赫", "")
                    reordered_data_rows.append(reordered_row)
                
                warranty_row_idx = -1
                warranty_years = 3
                for idx, r in enumerate(reordered_data_rows):
                    item_name = r[r_col_item].strip()
                    desc_text = r[r_col_desc].strip()
                    
                    match = None
                    if item_name:
                        match = re.search(r'^(\d+)年', item_name)
                    if not match:
                        match = re.search(r'^(\d+)年', desc_text)
                        
                    if match:
                        combined_text = (item_name + " " + desc_text).lower()
                        if any(kw in combined_text for kw in ("保固", "7x24", "小時", "到場", "維護", "維保", "教育訓練")):
                            warranty_row_idx = idx
                            warranty_years = int(match.group(1))
                            break
                
                original_ttl = sum(parse_price(r[r_col_total]) for r in reordered_data_rows)
                ratio = 0.35
                grand_total = original_ttl
                
                if warranty_row_idx != -1:
                    ratios = {1: 0.35, 2: 0.35, 3: 0.35, 4: 0.45, 5: 0.55, 6: 0.65}
                    ratio = ratios.get(warranty_years, 0.35)
                    grand_total = original_ttl + (original_ttl * ratio)
                
                is_new_version = any("項目" in h for h in headers)
                
                warranty_header_name = f"{warranty_years}年保固 單價(USD)/台"
                if is_new_version:
                    clean_headers = ["No", "項目", "項目內容", "Qty", "", "單價(USD)", warranty_header_name, "總計(USD)"]
                else:
                    clean_headers = ["No.", "專案", "內容", "Qty", "單位", "單價(USD)", warranty_header_name, "總計(USD)"]
                
                if has_remark_col:
                    clean_headers.append("備註")
                
                final_rows = [clean_headers]
                final_rows.extend(reordered_data_rows)
                
                if total_row_raw:
                    if is_new_version:
                        total_label = "Total"
                    else:
                        total_label = "TTL (USD)"
                    total_row = [""] * len(clean_headers)
                    total_row[r_col_warranty] = total_label
                    total_row[r_col_total] = f"{grand_total:,.0f}"
                    if has_remark_col and len(total_row_raw) > remark_col_idx:
                        total_row[len(clean_headers) - 1] = total_row_raw[remark_col_idx]
                    final_rows.append(total_row)
                    
                table_entry = {
                    "page": page_index,
                    "index": table_idx,
                    "title": f"Page {page_index} Table {table_idx}",
                    "rows": final_rows,
                    "top_y": top_y,
                    "bottom_y": bottom_y,
                    "customer_info": customer_info,
                    "remarks_lines": remarks_lines,
                    "warranty_years": warranty_years if warranty_row_idx != -1 else 3,
                    "is_new_version": is_new_version
                }
                if warranty_row_idx != -1:
                    table_entry["ratio"] = f"{ratio * 100:.0f}%"
                    table_entry["formula"] = f"TTL 計算公式: 原本的 TTL {original_ttl:,.0f} + 原本的 TTL {original_ttl:,.0f} * {ratio * 100:.0f}% = {grand_total:,.0f} (USD)"
                tables.append(table_entry)
                
    return tables


def merge_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return []

    merged_rows = []
    current_row = rows[0]

    for next_row in rows[1:]:
        has_content = any(cell.strip() for cell in next_row)
        is_continuation = (
            has_content and
            (not next_row[0].strip()) and
            (sum(1 for cell in next_row if not cell.strip()) > len(next_row) / 2)
        )

        if is_continuation:
            for col_idx in range(len(current_row)):
                if col_idx < len(next_row):
                    next_val = next_row[col_idx].strip()
                    if next_val:
                        if current_row[col_idx].strip():
                            current_row[col_idx] = current_row[col_idx] + " " + next_val
                        else:
                            current_row[col_idx] = next_val
        else:
            merged_rows.append(current_row)
            current_row = next_row

    merged_rows.append(current_row)
    return merged_rows


def normalize_table(table: list[list[object]]) -> list[list[str]]:
    rows = [[clean_cell(cell) for cell in row] for row in table]
    non_empty_rows = [row for row in rows if any(cell for cell in row)]
    return merge_rows(non_empty_rows)


def table_has_quote_data(rows: list[list[str]]) -> bool:
    joined = " ".join(" ".join(row) for row in rows).lower()
    quote_markers = ["product", "po price", "單價", "總計", "qty", "quotation", "nvidia", "beehe"]
    non_empty_cells = sum(1 for row in rows for cell in row if cell)
    return non_empty_cells >= 4 and any(marker in joined for marker in quote_markers)


def find_header_row_index(rows: list[list[str]]) -> int:
    for idx in range(1, len(rows)):
        first_cell = rows[idx][0].strip()
        if first_cell in ("1", "01", "1."):
            prev_cell = rows[idx - 1][0].strip()
            if not prev_cell.isdigit():
                return idx - 1

    header_keywords = {"no.", "no", "item", "product", "part number", "description", "專案", "內容", "單價"}
    for idx, row in enumerate(rows):
        for cell in row:
            val = cell.lower().strip()
            if any(kw in val for kw in header_keywords):
                return idx
    return 0


def is_footer_row(row: list[str]) -> bool:
    if not row:
        return True
    non_empty = [c.strip() for c in row if c.strip()]
    if not non_empty:
        return True
    if len(non_empty) == 1:
        val = non_empty[0].lower()
        footer_kws = ["備註", "remark", "note", "付款", "交易", "保固", "有效期限", "sign", "signature", "址", "tel", "fax"]
        if any(kw in val for kw in footer_kws) or len(val) > 40:
            return True
    return False


def clean_quote_table(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return []
    header_idx = find_header_row_index(rows)
    table_rows = rows[header_idx:]
    if not table_rows:
        return []

    cleaned_rows = [table_rows[0]]
    for row in table_rows[1:]:
        first_cell = row[0].strip()
        row_joined = " ".join(row).lower()

        if (first_cell.lower().startswith("ttl") or 
            first_cell.lower().startswith("total") or 
            re.search(r'\b(ttl|total)\b', row_joined)):
            joined_all = " ".join(c.strip() for c in row if c.strip())
            ttl_match = re.search(r"(?i)(ttl|total|grand\s*total)\s*(\([^)]+\))?\s*([\d,.]+)", joined_all)
            if ttl_match:
                new_row = [""] * len(row)
                label = ttl_match.group(1)
                if ttl_match.group(2):
                    label += " " + ttl_match.group(2)
                val = ttl_match.group(3)
                new_row[-1] = val
                label_idx = -2 if len(row) < 6 else -3
                new_row[label_idx] = label
                cleaned_rows.append(new_row)
            else:
                cleaned_rows.append(row)
            break

        if is_footer_row(row):
            break

        cleaned_rows.append(row)

    return cleaned_rows


def render_page_thumbnail(doc: object | None, page_index: int) -> str:
    if fitz is None or doc is None:
        return ""

    page = doc.load_page(page_index - 1)
    target_width = 1200.0
    scale = min(3.0, target_width / max(float(page.rect.width), 1.0))
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    encoded = base64.b64encode(pix.tobytes("jpg")).decode("ascii")
    return f"data:image/jpeg;base64,{encoded}"


def parse_amd_text_table(text: str) -> list[list[str]] | None:
    log("Running AMD fallback text-based table parser...")
    lines = [line.strip() for line in text.split("\n")]
    
    header_idx = -1
    for idx, line in enumerate(lines):
        if "Product" in line and "Start Date" in line and "End Date" in line and "Qty" in line and "Pricing Term" in line:
            header_idx = idx
            break
            
    if header_idx == -1:
        log("AMD fallback parser: Required header keywords not found in text.")
        return None
        
    log(f"AMD fallback parser: Found header keywords at line {header_idx}: '{lines[header_idx]}'")
    headers = ["Product", "Start Date", "End Date", "Qty", "Pricing Term"]
    rows = [headers]
    
    idx = header_idx + 1
    while idx < len(lines):
        line = lines[idx]
        if not line or line.startswith("___") or "Eligible Incentives" in line:
            log(f"AMD fallback parser: Stopping extraction at line {idx} due to stop delimiter: '{line}'")
            break
            
        date_matches = list(re.finditer(r"\b\d{2}/\d{2}/\d{4}\b", line))
        if len(date_matches) >= 2:
            start_date = date_matches[0].group(0)
            end_date = date_matches[1].group(0)
            product = line[:date_matches[0].start()].strip()
            
            rest = line[date_matches[1].end():].strip()
            rest_parts = rest.split()
            if len(rest_parts) >= 2:
                qty = rest_parts[0]
                price = rest_parts[1]
                term = " ".join(rest_parts[2:])
                
                if idx + 1 < len(lines):
                    next_line = lines[idx + 1].strip()
                    if next_line.isdigit() and len(next_line) == 1:
                        price = price + next_line
                        idx += 1
                        
                rows.append([product, start_date, end_date, qty, f"{price} {term}".strip()])
                log(f"AMD fallback parser: Extracted row: {rows[-1]}")
        idx += 1
        
    if len(rows) > 1:
        log(f"AMD fallback parser: Successfully extracted {len(rows) - 1} data rows.")
        return rows
    log("AMD fallback parser: No data rows were extracted.")
    return None


def extract_document(pdf_path: Path, target_pages: set[int] | None = None) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    log(f"Extracting tables from document: {pdf_path.name}")
    tables: list[dict[str, object]] = []
    pages: list[dict[str, object]] = []
    pdf_doc = fitz.open(pdf_path) if fitz is not None else None

    with pdfplumber.open(pdf_path) as pdf:
        is_beehe = False
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "beehe" in text.lower() or "比赫" in text:
                is_beehe = True
                break
                
        if is_beehe:
            beehe_tables = extract_beehe_document(pdf_path, target_pages)
            for page_index, page in enumerate(pdf.pages, start=1):
                if target_pages is not None and page_index not in target_pages:
                    continue
                text = page.extract_text() or ""
                pages.append({
                    "page": page_index,
                    "text": text,
                    "thumbnail": render_page_thumbnail(pdf_doc, page_index),
                })
            if pdf_doc is not None:
                pdf_doc.close()
            return beehe_tables, pages

        for page_index, page in enumerate(pdf.pages, start=1):
            if target_pages is not None and page_index not in target_pages:
                log(f"Page {page_index}: Skipped (not in target pages).")
                continue
            text = page.extract_text() or ""
            log(f"Page {page_index}: Extracted text (length={len(text)}).")
            pages.append(
                {
                    "page": page_index,
                    "text": text,
                    "thumbnail": render_page_thumbnail(pdf_doc, page_index),
                }
            )
            
            page_tables = []
            raw_tables = page.extract_tables() or []
            log(f"Page {page_index}: pdfplumber detected {len(raw_tables)} raw tables.")
            for table_idx, table in enumerate(raw_tables, start=1):
                normalized = normalize_table(table)
                if normalized and table_has_quote_data(normalized):
                    cleaned = clean_quote_table(normalized)
                    if cleaned and len(cleaned) > 1:
                        page_tables.append(cleaned)
                        log(f"Page {page_index} Table {table_idx}: Accepted grid table with {len(cleaned)} rows (including header).")
                    else:
                        log(f"Page {page_index} Table {table_idx}: Ignored grid table with {len(cleaned) if cleaned else 0} rows (no data).")
                else:
                    log(f"Page {page_index} Table {table_idx}: Ignored (does not match quote signature or too small).")
            
            if not page_tables and ("AMD" in text or "Advanced Micro Devices" in text):
                log(f"Page {page_index}: No grid tables accepted, but AMD detected. Triggering AMD fallback parser.")
                amd_table = parse_amd_text_table(text)
                if amd_table:
                    page_tables.append(amd_table)
            
            for table_idx, normalized in enumerate(page_tables, start=1):
                tables.append(
                    {
                        "page": page_index,
                        "index": table_idx,
                        "title": f"Page {page_index} Table {table_idx}",
                        "rows": normalized,
                    }
                )
                log(f"Page {page_index}: Registered table {table_idx} with {len(normalized)} rows.")
                
    if pdf_doc is not None:
        pdf_doc.close()
    return tables, pages


def write_table(ws, start_row: int, title: str, rows: list[list[str]]) -> int:
    max_cols = max((len(row) for row in rows), default=1)

    for row_offset, row in enumerate(rows):
        for col_index in range(1, max_cols + 1):
            cell = ws.cell(start_row + row_offset, col_index)
            cell.value = row[col_index - 1] if col_index <= len(row) else ""
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_offset == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E8DAC0")
    return start_row + len(rows) + 2


def autosize(ws) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 42)


def write_text_fallback(wb: Workbook, pdf_path: Path, target_pages: set[int] | None = None) -> None:
    ws = wb.create_sheet("Text")
    ws.append(["Page", "Text"])
    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            if target_pages is not None and page_index not in target_pages:
                continue
            ws.append([page_index, page.extract_text() or ""])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def convert_impl(pdf_path: Path, output_path: Path, target_pages: set[int] | None = None, project_name: str | None = None) -> None:
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    # Check if this was a Beehe PDF
    is_beehe = False
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "beehe" in text.lower() or "比赫" in text:
                is_beehe = True
                break

    tables, pages = extract_document(pdf_path, target_pages)
    wb = Workbook()
    ws = wb.active
    if project_name:
        sanitized_name = re.sub(r'[\\/*?:\[\]]', '_', project_name)[:31]
        ws.title = sanitized_name
    else:
        ws.title = "Tables"

    if is_beehe and fitz is not None:
        log("Applying custom Beehe layout with native styled cells and Logo crop.")
        pdf_doc = fitz.open(pdf_path)
        row_cursor = 2 if project_name else 1
        temp_images: list[Path] = []
        
        for table in tables:
            page_index = table["page"]
            page_fitz = pdf_doc.load_page(page_index - 1)
            width = page_fitz.rect.width
            height = page_fitz.rect.height
            
            # Crop Logo icon (from coords: x0=48.20, x1=92.32, top=65.14, bottom=106.14)
            logo_img_path = Path(tempfile.gettempdir()) / f"temp_{output_path.stem}_logo_{page_index}.png"
            pix_logo = page_fitz.get_pixmap(clip=fitz.Rect(45.0, 60.0, 95.0, 97.0), matrix=fitz.Matrix(3.0, 3.0))
            pix_logo.save(str(logo_img_path))
            temp_images.append(logo_img_path)
            
            # Write layout blocks
            customer_info = table.get("customer_info", {"customer": "", "contact": "", "phone": "", "date": "", "project": ""})
            remarks = table.get("remarks_lines", [])
            
            rem_header = "* 備註"
            rem_body = list(remarks)
            if rem_body and any(kw in rem_body[0] for kw in ("* 備註", "備註", "remark", "Remark")):
                rem_header = rem_body.pop(0)
                
            rows = table["rows"]
            max_cols = max((len(row) for row in rows), default=1)
            end_col_letter = get_column_letter(max_cols)
            is_new_version = table.get("is_new_version", False)
            
            warranty_years = table.get("warranty_years", 3)
            ratios = {1: 0.35, 2: 0.35, 3: 0.35, 4: 0.45, 5: 0.55, 6: 0.65}
            ratio = ratios.get(warranty_years, 0.35)
            thin_side = Side(style="thin", color="000000")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            start_row = row_cursor + 8
            table.setdefault("table_start_row", start_row)
            
            # The last row of remarks will be at:
            end_row = start_row + len(rows) + 2 + len(rem_body)
            
            # Apply solid white background to the entire page block (A{row_cursor} to end_col_letter{end_row})
            white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            style_range(ws, f"A{row_cursor}:{end_col_letter}{end_row}", fill=white_fill)
            
            # 1. Company Header Area
            ws.row_dimensions[row_cursor].height = 18
            ws.row_dimensions[row_cursor + 1].height = 16
            ws.row_dimensions[row_cursor + 2].height = 16
            ws.row_dimensions[row_cursor + 3].height = 25
            
            ws.merge_cells(f"A{row_cursor}:B{row_cursor+2}")
            ws.merge_cells(f"C{row_cursor}:{end_col_letter}{row_cursor}")
            ws.merge_cells(f"C{row_cursor+1}:{end_col_letter}{row_cursor+1}")
            ws.merge_cells(f"C{row_cursor+2}:{end_col_letter}{row_cursor+2}")
            ws.merge_cells(f"A{row_cursor+3}:{end_col_letter}{row_cursor+3}")
            
            ws.cell(row_cursor, 3).value = "Beehe Electric (Taicang) Co., Ltd"
            ws.cell(row_cursor + 1, 3).value = "ADD: No.5-1 Shuanghu Road, Taicang City, Jiangsu Province,China"
            ws.cell(row_cursor + 2, 3).value = "TEL：0086-512-53983090 etx.8910"
            ws.cell(row_cursor + 3, 1).value = "Quotation"
            
            # Style text
            style_range(ws, f"C{row_cursor}:{end_col_letter}{row_cursor}", font=Font(name="Calibri", size=12, bold=True), alignment=Alignment(vertical="center", horizontal="left"))
            style_range(ws, f"C{row_cursor+1}:{end_col_letter}{row_cursor+2}", font=Font(name="Calibri", size=9), alignment=Alignment(vertical="center", horizontal="left"))
            style_range(ws, f"A{row_cursor+3}:{end_col_letter}{row_cursor+3}", font=Font(name="Calibri", size=16, bold=True), alignment=Alignment(vertical="center", horizontal="center"))
            
            # Insert logo image in A{row_cursor} (spans A1:B3)
            img_logo = Image(str(logo_img_path))
            img_logo.width = int(50 * 96 / 72)
            img_logo.height = int(50 * 96 / 72)
            
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            # Shift 2 pixels right (19050 EMUs) and 2 pixels down (19050 EMUs) to avoid border overlap
            marker = AnchorMarker(col=0, colOff=19050, row=row_cursor-1, rowOff=19050)
            ext = XDRPositiveSize2D(cx=img_logo.width * 9525, cy=img_logo.height * 9525)
            img_logo.anchor = OneCellAnchor(_from=marker, ext=ext)
            ws.add_image(img_logo)
            
            # 2. Customer Info Area (Row 5 to Row 8)
            ws.row_dimensions[row_cursor + 4].height = 18
            ws.row_dimensions[row_cursor + 5].height = 18
            ws.row_dimensions[row_cursor + 6].height = 18
            ws.row_dimensions[row_cursor + 7].height = 18
            
            raw_text = customer_info.get("raw_text", "")
            cust_label = "客户 : " if "客户" in raw_text else "客戶 : "
            ws.cell(row_cursor + 4, 1).value = cust_label + customer_info["customer"]
            
            contact_label = "聯繫 : " if "聯繫" in raw_text or "專案" in raw_text else "联系 : "
            ws.cell(row_cursor + 5, 1).value = contact_label + customer_info["contact"]
            
            date_label = "日期 : " if "日期 :" in raw_text and "報價日期" not in raw_text else "報價日期 : "
            ws.cell(row_cursor + 5, 6).value = date_label + customer_info["date"]
            
            if customer_info.get("mail"):
                ws.cell(row_cursor + 6, 1).value = "Mail : " + customer_info["mail"]
            else:
                ws.cell(row_cursor + 6, 1).value = "電話 : " + customer_info["phone"]
                
            ws.cell(row_cursor + 7, 1).value = ("專案 : " + customer_info["project"]) if customer_info.get("project") else ""
            
            # Apply styles and borders before merging to ensure openpyxl serializes outer borders correctly
            for r in range(row_cursor + 4, row_cursor + 8):
                for c in range(1, max_cols + 1):
                    cell = ws.cell(r, c)
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
                    
            ws.merge_cells(start_row=row_cursor+5, start_column=1, end_row=row_cursor+5, end_column=5)
            ws.merge_cells(start_row=row_cursor+5, start_column=6, end_row=row_cursor+5, end_column=max_cols)
            ws.merge_cells(start_row=row_cursor+6, start_column=1, end_row=row_cursor+6, end_column=max_cols)
            ws.merge_cells(start_row=row_cursor+7, start_column=1, end_row=row_cursor+7, end_column=max_cols)
            
            # Apply outer border around the entire company header + customer info block (Rows 1-8) with divider line
            apply_outer_borders(ws, row_cursor, row_cursor + 7, 1, max_cols, divider_row=row_cursor + 3)
            
            # 3. Write Quotation Table starting at row_cursor + 8 (Row 9)
            ws.row_dimensions[start_row].height = 28
            max_cols = max((len(row) for row in rows), default=1)
            for row_offset, row in enumerate(rows):
                r_idx = start_row + row_offset
                for col_index in range(1, max_cols + 1):
                    cell = ws.cell(r_idx, col_index)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = thin_border
                    
                    val_str = str(row[col_index - 1]).strip() if col_index <= len(row) else ""
                    
                    if row_offset == 0:
                        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                        cell.fill = PatternFill("solid", fgColor="D9E1F2")
                        if col_index == 7: # Column G is always the warranty column header
                            red_font = InlineFont(rFont="Calibri", sz=11, b=True, color="FF0000")
                            black_font = InlineFont(rFont="Calibri", sz=11, b=True, color="000000")
                            cell.value = CellRichText(
                                TextBlock(red_font, f"{warranty_years}年保固\n"),
                                TextBlock(black_font, "單價(USD)/台")
                            )
                        else:
                            cell.value = val_str
                            cell.font = Font(bold=True)
                    elif row_offset == len(rows) - 1: # Total row
                        if col_index == 8: # Column H is always the total sum column
                            cell.value = f"=SUM(H{start_row+1}:H{start_row+len(rows)-2}) * (1 + {ratio})"
                            cell.number_format = "#,##0"
                        else:
                            cell.value = val_str
                    else: # Data rows (same column layout for both versions)
                        if col_index == 4: # Qty
                            try:
                                cell.value = int(val_str.replace(",", ""))
                            except ValueError:
                                cell.value = val_str
                        elif col_index == 5: # Unit / Empty
                            cell.value = val_str
                            cell.alignment = Alignment(vertical="center", horizontal="center")
                        elif col_index in (6, 8): # Price, Total
                            try:
                                cell.value = float(val_str.replace(",", ""))
                                cell.number_format = "#,##0"
                            except ValueError:
                                cell.value = val_str
                        elif col_index == 7: # Warranty
                            if val_str == "-":
                                cell.value = "-"
                            else:
                                try:
                                    cell.value = float(val_str.replace(",", ""))
                                    cell.number_format = "#,##0"
                                except ValueError:
                                    cell.value = val_str
                        else:
                            cell.value = val_str
            
            if is_new_version:
                ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=5)
            
            # Write new calculation table on the right (starting at max_cols + 2, aligning with row_cursor)
            calc_start_row = row_cursor
            calc_start_col = max_cols + 2
            new_table_headers = ["年", "金額", "Percentage", "金額"]
            new_table_cols = [calc_start_col, calc_start_col + 1, calc_start_col + 2, calc_start_col + 3]
            
            # Letters for the calculation columns
            col_year_letter = get_column_letter(calc_start_col)
            col_orig_letter = get_column_letter(calc_start_col + 1)
            col_pct_letter = get_column_letter(calc_start_col + 2)
            col_calc_letter = get_column_letter(calc_start_col + 3)
            
            # Write headers
            for c_offset, h_text in enumerate(new_table_headers):
                c_idx = new_table_cols[c_offset]
                cell = ws.cell(calc_start_row, c_idx)
                cell.value = h_text
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.border = thin_border
            
            # Ratios for years 1 to 8
            year_ratios = {
                1: 0.35,
                2: 0.35,
                3: 0.35,
                4: 0.45,
                5: 0.55,
                6: 0.65,
                7: 0.75,
                8: 0.85
            }
            
            for year in range(1, 9):
                r = calc_start_row + year
                
                # Col 1: Year
                cell_j = ws.cell(r, calc_start_col)
                cell_j.value = year
                cell_j.font = Font(name="Calibri", size=10)
                cell_j.alignment = Alignment(vertical="center", horizontal="center")
                cell_j.border = thin_border
                
                # Col 2: Original TTL (金額)
                cell_k = ws.cell(r, calc_start_col + 1)
                cell_k.value = f'=IF({col_year_letter}{r}={warranty_years}, SUM($H${start_row+1}:$H${start_row+len(rows)-2}), "")'
                cell_k.font = Font(name="Calibri", size=10)
                cell_k.alignment = Alignment(vertical="center", horizontal="right")
                cell_k.number_format = "$#,##0"
                cell_k.border = thin_border
                
                # Col 3: Percentage
                cell_l = ws.cell(r, calc_start_col + 2)
                cell_l.value = year_ratios[year]
                cell_l.font = Font(name="Calibri", size=10)
                cell_l.alignment = Alignment(vertical="center", horizontal="right")
                cell_l.number_format = "0%"
                cell_l.border = thin_border
                
                # Col 4: Calculated TTL (金額)
                cell_m = ws.cell(r, calc_start_col + 3)
                cell_m.value = f'=IF({col_orig_letter}{r}<>"", {col_orig_letter}{r}*(1+{col_pct_letter}{r}), 0)'
                cell_m.font = Font(name="Calibri", size=10)
                cell_m.alignment = Alignment(vertical="center", horizontal="right")
                cell_m.number_format = "$#,##0"
                cell_m.border = thin_border
                        
            # Apply blackout to data rows in this table
            black_fill = PatternFill(fill_type="solid", fgColor="000000")
            black_font = Font(color="000000")
            
            blackout_cols = []
            for c_idx in range(1, max_cols + 1):
                col_header = str(ws.cell(start_row, c_idx).value or "").strip().lower()
                if any(kw in col_header for kw in ("單價", "price", "保固", "warranty", "總計", "total", "ttl")) and "內容" not in col_header and "備註" not in col_header and "remark" not in col_header:
                    blackout_cols.append(c_idx)
                    
            curr_row = start_row + 1
            while curr_row < start_row + len(rows) - 1: # exclude header and total row
                for c_idx in blackout_cols:
                    ws.cell(curr_row, c_idx).fill = black_fill
                    ws.cell(curr_row, c_idx).font = black_font
                curr_row += 1
                
            # 4. Remarks Section (A18 onwards)
            if remarks:
                end_col_letter = get_column_letter(max_cols)
                ws.row_dimensions[curr_row + 1].height = 18
                ws.cell(curr_row + 1, 1).value = rem_header
                ws.cell(curr_row + 1, 1).font = Font(name="Calibri", size=10, bold=True)
                ws.merge_cells(start_row=curr_row+1, start_column=1, end_row=curr_row+1, end_column=max_cols)
                style_range(ws, f"A{curr_row+1}:{end_col_letter}{curr_row+1}", alignment=Alignment(vertical="center", horizontal="left", indent=1))
                
                r_rem = curr_row + 2
                for line in rem_body:
                    ws.row_dimensions[r_rem].height = 18
                    ws.cell(r_rem, 1).value = line
                    ws.merge_cells(start_row=r_rem, start_column=1, end_row=r_rem, end_column=max_cols)
                    style_range(ws, f"A{r_rem}:{end_col_letter}{r_rem}", font=Font(name="Calibri", size=9), alignment=Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1))
                    r_rem += 1
                    
                # Apply outer border around the Remarks section block
                apply_outer_borders(ws, curr_row + 1, r_rem - 1, 1, max_cols)
                    
                # Spacer row
                ws.row_dimensions[r_rem].height = 25
                row_cursor = r_rem + 1
            else:
                row_cursor = curr_row + 1
            
        pdf_doc.close()
        
        # Auto fit columns (excluding image/remarks rows width calculations)
        max_cols_all = max((len(t["rows"][0]) for t in tables), default=8)
        separator_col = max_cols_all + 1
        calc_start_col = max_cols_all + 2
        
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            
            if col_idx == separator_col: # Separator column, make it narrow
                ws.column_dimensions[col_letter].width = 3
                continue
                
            if col_idx in (calc_start_col + 1, calc_start_col + 3): # Fixed width columns in calculation table
                ws.column_dimensions[col_letter].width = 12
                continue
                
            max_len = 0
            for table in tables:
                t_start = table.get("table_start_row", 10 if project_name else 9)
                t_end = t_start + len(table["rows"])
                if col_idx >= calc_start_col:
                    r_start = t_start - 8
                    r_end = t_start
                else:
                    r_start = t_start
                    r_end = t_end
                for r_idx in range(r_start, r_end + 1):
                    val = str(ws.cell(r_idx, col_idx).value or "")
                    if val.startswith("="):
                        continue
                    if val:
                        line_lens = [get_display_length(line) for line in val.split("\n")]
                        max_line_len = max(line_lens, default=0)
                        if max_line_len > max_len:
                            max_len = max_line_len
                            
            if col_idx >= calc_start_col:
                min_w = 6 if col_idx == calc_start_col else 10
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), 25)
            else:
                ws.column_dimensions[col_letter].width = min(max(max_len + 5, 10), 42)
            
        ws.freeze_panes = None
        
    else:
        row_cursor = 2 if project_name else 1
        for table in tables:
            row_cursor = write_table(ws, row_cursor, table["title"], table["rows"])
        if not tables:
            ws.append(["No table-like quotation data detected. See Text sheet."])
        autosize(ws)
        ws.freeze_panes = "A3" if project_name else "A2"

    if project_name:
        ws.cell(1, 1).value = project_name
        ws.cell(1, 1).font = Font(name="Calibri", size=24, bold=True)
        ws.cell(1, 1).alignment = Alignment(vertical="center", horizontal="center")
        ws.row_dimensions[1].height = 35
        
        # Merge A1 across max_cols_all
        max_cols_all = max((len(t["rows"][0]) for t in tables), default=8)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols_all)

    write_text_fallback(wb, pdf_path, target_pages)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    if project_name:
        output_dir = output_path.parent
        latest_master = find_latest_master_path(output_dir)
        try:
            from openpyxl import load_workbook
            if latest_master and latest_master.exists():
                wb_master = load_workbook(latest_master)
            else:
                wb_master = Workbook()
                if "Sheet" in wb_master.sheetnames:
                    wb_master.remove(wb_master["Sheet"])
            
            sanitized_name = re.sub(r'[\\/*?:\[\]]', '_', project_name)[:31]
            copy_sheet_to_workbook(ws, wb_master, sanitized_name)
            
            # Save to a new filename with timestamp
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            new_master_path = output_dir / f"專案匯總管理表_{timestamp}.xlsx"
            
            wb_master.save(new_master_path)
            wb_master.close()
        except Exception as e:
            log(f"Error copying sheet to master workbook: {e}")

    # Clean up temp images
    if is_beehe:
        for p in temp_images:
            try:
                p.unlink()
            except Exception:
                pass

    payload = {
        "input": str(pdf_path),
        "output": str(output_path),
        "table_count": len(tables),
        "tables": tables,
        "pages": pages,
        "logs": LOGS,
    }
    print(json.dumps(payload, ensure_ascii=False))


def convert(pdf_path: Path, output_path: Path, target_pages: set[int] | None = None, project_name: str | None = None) -> None:
    try:
        convert_impl(pdf_path, output_path, target_pages, project_name)
    except Exception as error:
        payload = {
            "input": str(pdf_path),
            "output": str(output_path),
            "status": "error",
            "message": str(error),
            "table_count": 0,
            "tables": [],
            "pages": [],
            "logs": LOGS + [f"Critical error: {error}"],
        }
        print(json.dumps(payload, ensure_ascii=False))


def get_thumbnails_impl(pdf_path: Path) -> None:
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    pages = []
    pdf_doc = fitz.open(pdf_path) if fitz is not None else None
    if pdf_doc is not None:
        for page_index in range(1, len(pdf_doc) + 1):
            pages.append(
                {
                    "page": page_index,
                    "text": "",
                    "thumbnail": render_page_thumbnail(pdf_doc, page_index),
                }
            )
        pdf_doc.close()
    else:
        # Fallback if fitz is not available
        with pdfplumber.open(pdf_path) as pdf:
            for page_index in range(1, len(pdf.pages) + 1):
                pages.append(
                    {
                        "page": page_index,
                        "text": "",
                        "thumbnail": "",
                    }
                )

    payload = {
        "input": str(pdf_path),
        "output": "",
        "table_count": 0,
        "tables": [],
        "pages": pages,
        "logs": LOGS,
    }
    print(json.dumps(payload, ensure_ascii=False))


def get_thumbnails(pdf_path: Path) -> None:
    try:
        get_thumbnails_impl(pdf_path)
    except Exception as error:
        payload = {
            "input": str(pdf_path),
            "output": "",
            "status": "error",
            "message": str(error),
            "table_count": 0,
            "tables": [],
            "pages": [],
            "logs": LOGS + [f"Critical error: {error}"],
        }
        print(json.dumps(payload, ensure_ascii=False))


def run_daemon(port: int) -> None:
    import socket
    import io
    import contextlib

    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    server.bind(("127.0.0.1", port))
    server.listen(5)
    log(f"Daemon socket server listening on 127.0.0.1:{port}")
    
    # Notify Tauri that daemon is ready
    print(json.dumps({"status": "ready"}))
    sys.stdout.flush()
    
    while True:
        try:
            conn, addr = server.accept()
            data = b""
            while b"\n" not in data:
                chunk = conn.recv(4096)
                if not chunk:
                    break
                data += chunk
            if not data:
                conn.close()
                continue
                
            line = data.decode("utf-8").strip()
            if not line:
                conn.close()
                continue
                
            cmd = json.loads(line)
            action = cmd.get("action")
            
            if action == "shutdown":
                conn.sendall(json.dumps({"status": "shutdown"}).encode("utf-8") + b"\n")
                conn.close()
                break
                
            elif action == "convert":
                pdf_path = Path(cmd["pdf_path"])
                output_path = Path(cmd["output_path"])
                pages = cmd.get("pages")
                target_pages = set(pages) if pages else None
                project_name = cmd.get("project_name")
                
                f = io.StringIO()
                with contextlib.redirect_stdout(f):
                    convert(pdf_path, output_path, target_pages, project_name)
                response = f.getvalue().strip()
                if not response:
                    response = json.dumps({"status": "error", "message": "Empty response from converter"})
                conn.sendall(response.encode("utf-8") + b"\n")
                
            elif action == "get_thumbnails":
                pdf_path = Path(cmd["pdf_path"])
                
                f = io.StringIO()
                with contextlib.redirect_stdout(f):
                    get_thumbnails(pdf_path)
                response = f.getvalue().strip()
                if not response:
                    response = json.dumps({"status": "error", "message": "Empty response from thumbnail extractor"})
                conn.sendall(response.encode("utf-8") + b"\n")
                
            elif action == "get_master_sheets":
                output_dir = Path(cmd["output_dir"])
                latest_master = find_latest_master_path(output_dir)
                sheets = []
                if latest_master and latest_master.exists():
                    try:
                        from openpyxl import load_workbook
                        wb_master = load_workbook(latest_master, read_only=True, keep_links=False)
                        sheets = wb_master.sheetnames
                        wb_master.close()
                    except Exception as e:
                        log(f"Error reading master workbook sheets: {e}")
                conn.sendall(json.dumps({"status": "success", "sheets": sheets}).encode("utf-8") + b"\n")
                
            else:
                conn.sendall(json.dumps({"status": "error", "message": f"Unknown action: {action}"}).encode("utf-8") + b"\n")
                
            conn.close()
        except Exception as e:
            log(f"Daemon connection error: {e}")
            try:
                conn.sendall(json.dumps({"status": "error", "message": str(e)}).encode("utf-8") + b"\n")
                conn.close()
            except Exception:
                pass


def main() -> int:
    if "--daemon" in sys.argv:
        try:
            idx = sys.argv.index("--daemon")
            port = int(sys.argv[idx + 1])
            run_daemon(port)
            return 0
        except Exception as e:
            print(f"Error starting daemon: {e}", file=sys.stderr)
            return 1

    if len(sys.argv) < 3:
        print("Usage: convert_pdf.py <input.pdf> <output.xlsx | --thumbnails-only> [--pages 1,2,3] [--project-name 'Name']", file=sys.stderr)
        return 2
    try:
        input_path = Path(sys.argv[1])
        output_arg = sys.argv[2]
        
        target_pages = None
        if "--pages" in sys.argv:
            try:
                pages_idx = sys.argv.index("--pages")
                pages_str = sys.argv[pages_idx + 1]
                target_pages = {int(p) for p in pages_str.split(",") if p.strip()}
            except Exception as e:
                print(f"Error parsing --pages: {e}", file=sys.stderr)
                return 2

        project_name = None
        if "--project-name" in sys.argv:
            try:
                project_name_idx = sys.argv.index("--project-name")
                project_name = sys.argv[project_name_idx + 1]
            except Exception as e:
                print(f"Error parsing --project-name: {e}", file=sys.stderr)
                return 2

        if output_arg == "--thumbnails-only":
            get_thumbnails(input_path)
        else:
            convert(input_path, Path(output_arg), target_pages, project_name)
    except Exception as error:
        print(str(error), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
